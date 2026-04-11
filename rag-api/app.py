import hashlib
import json
import csv
import logging
import os
import re
import subprocess
import threading
import time
import uuid
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any, Iterable, Iterator

from docx import Document
from fastapi import FastAPI, HTTPException
from openpyxl import load_workbook
from pydantic import BaseModel
from pypdf import PdfReader
from qdrant_client import QdrantClient
from qdrant_client.http import models as qmodels
from openai import OpenAI
from starlette.responses import StreamingResponse

try:
    import fitz
except Exception:
    fitz = None

try:
    from sentence_transformers import SentenceTransformer
except Exception:
    SentenceTransformer = None


def env_int(name: str, default: int) -> int:
    value = os.getenv(name)
    return int(value) if value else default


def env_bool(name: str, default: bool = False) -> bool:
    value = os.getenv(name)
    if value is None:
        return default
    return value.strip().lower() in {"1", "true", "yes", "on"}


KB_ROOT = Path(os.getenv("KB_ROOT", "/kb")).resolve()
KB_EXTENSIONS = tuple(
    ext.strip().lower()
    for ext in os.getenv("KB_EXTENSIONS", ".txt,.md,.markdown,.pdf,.docx").split(",")
    if ext.strip()
)
KB_EXCLUDE_PATTERNS = [
    value.strip() for value in os.getenv("KB_EXCLUDE_PATTERNS", "").split(",") if value.strip()
]
INDEX_BATCH_SIZE = env_int("INDEX_BATCH_SIZE", 32)
EMBEDDING_BATCH_SIZE = env_int("EMBEDDING_BATCH_SIZE", 128)
QDRANT_UPSERT_RETRIES = env_int("QDRANT_UPSERT_RETRIES", 4)
QDRANT_UPSERT_RETRY_DELAY_SECONDS = float(os.getenv("QDRANT_UPSERT_RETRY_DELAY_SECONDS", "2.0"))
CHUNK_SIZE = env_int("CHUNK_SIZE", 700)
CHUNK_OVERLAP = env_int("CHUNK_OVERLAP", 100)
RETRIEVAL_LIMIT = env_int("RETRIEVAL_LIMIT", 4)
MIN_RETRIEVAL_SCORE = float(os.getenv("MIN_RETRIEVAL_SCORE", "0.35"))
MAX_CONTEXT_CHARS = env_int("MAX_CONTEXT_CHARS", 1200)
MAX_OUTPUT_TOKENS = env_int("MAX_OUTPUT_TOKENS", 384)
QDRANT_URL = os.getenv("QDRANT_URL", "http://qdrant:6333")
QDRANT_COLLECTION = os.getenv("QDRANT_COLLECTION", "kodbox_kb")
OPENAI_BASE_URL = os.getenv("OPENAI_BASE_URL", "")
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY", "")
CHAT_MODEL = os.getenv("CHAT_MODEL", "")
EMBEDDING_PROVIDER = os.getenv("EMBEDDING_PROVIDER", "local").lower()
EMBEDDING_MODEL = os.getenv("EMBEDDING_MODEL", "BAAI/bge-small-zh-v1.5")
EMBEDDING_BASE_URL = os.getenv("EMBEDDING_BASE_URL", OPENAI_BASE_URL)
EMBEDDING_API_KEY = os.getenv("EMBEDDING_API_KEY", OPENAI_API_KEY)
EMBEDDING_MODEL_REMOTE = os.getenv("EMBEDDING_MODEL_REMOTE", "text-embedding-3-small")
PDF_EXTRACTOR = os.getenv("PDF_EXTRACTOR", "pymupdf").lower()
EXCEL_MAX_ROWS_PER_SHEET = env_int("EXCEL_MAX_ROWS_PER_SHEET", 5000)
EXCEL_MAX_COLUMNS_PER_ROW = env_int("EXCEL_MAX_COLUMNS_PER_ROW", 200)
EXCEL_INCLUDE_EMPTY_ROWS = os.getenv("EXCEL_INCLUDE_EMPTY_ROWS", "false").lower() == "true"
STATE_DIR = Path(os.getenv("STATE_DIR", "/var/lib/rag-api")).resolve()
MANIFEST_PATH = STATE_DIR / "index_manifest.json"
JOBS_DIR = STATE_DIR / "jobs"
DEBUG = env_bool("DEBUG", False)

app = FastAPI(title="lightweight-rag-api", version="0.1.0")
logger = logging.getLogger("rag-api")
if not logging.getLogger().handlers:
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(name)s %(message)s")
logger.setLevel(logging.INFO)


@dataclass
class Chunk:
    id: str
    text: str
    payload: dict[str, Any]


class ChatMessage(BaseModel):
    role: str
    content: str | list[dict[str, Any]] | None = None


class ChatCompletionRequest(BaseModel):
    model: str | None = None
    messages: list[ChatMessage]
    temperature: float | None = 0.2
    stream: bool | None = False


class IndexResponse(BaseModel):
    status: str = "completed"
    job_id: str | None = None
    indexed_files: int
    deleted_files: int
    indexed_chunks: int
    scanned_files: int = 0
    skipped_files: int = 0
    started_at: float | None = None
    finished_at: float | None = None
    timings: dict[str, float] = {}


class IndexJobStatus(BaseModel):
    job_id: str
    status: str
    started_at: float | None = None
    finished_at: float | None = None
    error: str | None = None
    progress: dict[str, Any] | None = None
    result: IndexResponse | None = None


class EmbeddingService:
    def __init__(self) -> None:
        self._local_model = None
        self._remote_client = None

    def embed(self, texts: list[str]) -> list[list[float]]:
        if not texts:
            return []
        if EMBEDDING_PROVIDER == "openai":
            return self._embed_remote(texts)
        return self._embed_local(texts)

    def dimension(self) -> int:
        probe = self.embed(["dimension probe"])
        return len(probe[0])

    def _embed_local(self, texts: list[str]) -> list[list[float]]:
        if SentenceTransformer is None:
            raise RuntimeError("sentence-transformers is not available")
        if self._local_model is None:
            self._local_model = SentenceTransformer(EMBEDDING_MODEL)
        vectors = self._local_model.encode(
            texts,
            batch_size=EMBEDDING_BATCH_SIZE,
            normalize_embeddings=True,
            show_progress_bar=False,
            convert_to_numpy=True,
        )
        return [vector.tolist() for vector in vectors]

    def _embed_remote(self, texts: list[str]) -> list[list[float]]:
        if self._remote_client is None:
            if not EMBEDDING_BASE_URL or not EMBEDDING_API_KEY:
                raise RuntimeError("remote embeddings require EMBEDDING_BASE_URL and EMBEDDING_API_KEY")
            self._remote_client = OpenAI(base_url=EMBEDDING_BASE_URL, api_key=EMBEDDING_API_KEY)
        response = self._remote_client.embeddings.create(model=EMBEDDING_MODEL_REMOTE, input=texts)
        return [item.embedding for item in response.data]


class ChatService:
    def __init__(self) -> None:
        if not OPENAI_BASE_URL or not OPENAI_API_KEY or not CHAT_MODEL:
            raise RuntimeError("chat model configuration is incomplete")
        self.client = OpenAI(base_url=OPENAI_BASE_URL, api_key=OPENAI_API_KEY)

    def complete(self, messages: list[dict[str, Any]], temperature: float) -> str:
        response = self.client.chat.completions.create(
            model=CHAT_MODEL,
            messages=messages,
            temperature=temperature,
            max_tokens=MAX_OUTPUT_TOKENS,
        )
        return response.choices[0].message.content or ""

    def complete_stream(self, messages: list[dict[str, Any]], temperature: float):
        return self.client.chat.completions.create(
            model=CHAT_MODEL,
            messages=messages,
            temperature=temperature,
            max_tokens=MAX_OUTPUT_TOKENS,
            stream=True,
        )


class VectorStore:
    def __init__(self, embedder: EmbeddingService) -> None:
        self.client = QdrantClient(url=QDRANT_URL)
        self.embedder = embedder
        self._ensure_collection()

    def _ensure_collection(self) -> None:
        existing = [collection.name for collection in self.client.get_collections().collections]
        if QDRANT_COLLECTION in existing:
            return
        self.client.create_collection(
            collection_name=QDRANT_COLLECTION,
            vectors_config=qmodels.VectorParams(
                size=self.embedder.dimension(),
                distance=qmodels.Distance.COSINE,
            ),
        )

    def build_points(self, chunks: list[Chunk]) -> list[qmodels.PointStruct]:
        if not chunks:
            return []
        points = []
        texts = [chunk.text for chunk in chunks]
        vectors = self.embedder.embed(texts)
        for chunk, vector in zip(chunks, vectors, strict=True):
            points.append(
                qmodels.PointStruct(
                    id=chunk.id,
                    vector=vector,
                    payload={**chunk.payload, "text": chunk.text},
                )
            )
        return points

    def write_points(self, points: list[qmodels.PointStruct]) -> None:
        if not points:
            return
        retry_delay = QDRANT_UPSERT_RETRY_DELAY_SECONDS
        last_error = None
        for attempt in range(1, QDRANT_UPSERT_RETRIES + 1):
            try:
                self.client.upsert(collection_name=QDRANT_COLLECTION, points=points)
                return
            except Exception as exc:
                last_error = exc
                if attempt >= QDRANT_UPSERT_RETRIES:
                    break
                time.sleep(retry_delay)
                retry_delay *= 2
        if last_error is not None:
            raise last_error

    def upsert_chunks(self, chunks: list[Chunk]) -> None:
        self.write_points(self.build_points(chunks))

    def delete_file(self, file_path: str) -> None:
        self.client.delete(
            collection_name=QDRANT_COLLECTION,
            points_selector=qmodels.FilterSelector(
                filter=qmodels.Filter(
                    must=[qmodels.FieldCondition(key="source_path", match=qmodels.MatchValue(value=file_path))]
                )
            ),
        )

    def list_indexed_files(self) -> dict[str, dict[str, Any]]:
        records: dict[str, dict[str, Any]] = {}
        next_offset = None
        while True:
            points, next_offset = self.client.scroll(
                collection_name=QDRANT_COLLECTION,
                limit=256,
                offset=next_offset,
                with_payload=True,
                with_vectors=False,
            )
            for point in points:
                payload = point.payload or {}
                source_path = payload.get("source_path")
                if not source_path or source_path in records:
                    continue
                records[source_path] = {
                    "file_hash": payload.get("file_hash"),
                    "mtime": payload.get("mtime"),
                }
            if next_offset is None:
                break
        return records

    def search(self, query: str, limit: int) -> list[dict[str, Any]]:
        vector = self.embedder.embed([query])[0]
        results = self.client.query_points(
            collection_name=QDRANT_COLLECTION,
            query=vector,
            limit=limit,
            with_payload=True,
            with_vectors=False,
        )
        matches = []
        for point in results.points:
            payload = point.payload or {}
            matches.append(
                {
                    "score": point.score,
                    "text": payload.get("text", ""),
                    "source_path": payload.get("source_path", ""),
                    "source_ref": payload.get("source_ref", payload.get("source_path", "")),
                    "file_name": payload.get("file_name", ""),
                    "sheet_name": payload.get("sheet_name", ""),
                    "chunk_index": payload.get("chunk_index", 0),
                }
            )
        return matches

    def search_filename_matches(self, query: str, limit: int) -> list[dict[str, Any]]:
        query_key = normalize_lookup_key(query)
        if not query_key:
            return []

        matches = []
        seen = set()
        next_offset = None
        while len(matches) < limit:
            points, next_offset = self.client.scroll(
                collection_name=QDRANT_COLLECTION,
                limit=256,
                offset=next_offset,
                with_payload=True,
                with_vectors=False,
            )
            if not points:
                break

            for point in points:
                payload = point.payload or {}
                source_ref = payload.get("source_ref", payload.get("source_path", ""))
                chunk_index = payload.get("chunk_index", 0)
                if not source_ref:
                    continue

                score = filename_match_score(query_key, payload)
                if score <= 0:
                    continue

                key = (source_ref, chunk_index)
                if key in seen:
                    continue
                seen.add(key)
                matches.append(
                    {
                        "score": score,
                        "text": payload.get("text", ""),
                        "source_path": payload.get("source_path", ""),
                        "source_ref": source_ref,
                        "file_name": payload.get("file_name", ""),
                        "sheet_name": payload.get("sheet_name", ""),
                        "chunk_index": chunk_index,
                    }
                )
                if len(matches) >= limit:
                    break

            if next_offset is None:
                break

        matches.sort(key=lambda item: item.get("score", 0.0), reverse=True)
        return matches[:limit]


class FileIndexer:
    def __init__(self, store: VectorStore) -> None:
        self.store = store

    def index(self, progress_callback: Any = None) -> IndexResponse:
        started_at = time.time()
        total_started = time.perf_counter()
        timings = init_timings()

        scan_started = time.perf_counter()
        current_files = list(self.iter_files())
        timings["scan_seconds"] += time.perf_counter() - scan_started

        manifest_started = time.perf_counter()
        manifest = load_manifest()
        timings["manifest_load_seconds"] += time.perf_counter() - manifest_started

        current_paths = {str(path) for path in current_files}
        deleted_files = 0
        indexed_files = 0
        indexed_chunks = 0
        skipped_files = 0

        report_index_progress(
            progress_callback,
            phase="scan",
            scanned_files=len(current_files),
            indexed_files=indexed_files,
            deleted_files=deleted_files,
            indexed_chunks=indexed_chunks,
            skipped_files=skipped_files,
            current_file=None,
            timings=round_timings(timings),
        )

        for source_path in set(manifest) - current_paths:
            delete_started = time.perf_counter()
            self.store.delete_file(source_path)
            timings["delete_seconds"] += time.perf_counter() - delete_started
            manifest.pop(source_path, None)
            deleted_files += 1

            report_index_progress(
                progress_callback,
                phase="delete",
                scanned_files=len(current_files),
                indexed_files=indexed_files,
                deleted_files=deleted_files,
                indexed_chunks=indexed_chunks,
                skipped_files=skipped_files,
                current_file=source_path,
                timings=round_timings(timings),
            )

        for file_number, path in enumerate(current_files, start=1):
            stat_started = time.perf_counter()
            stats = path.stat()
            path_key = str(path)
            mtime = stats.st_mtime
            size = stats.st_size
            timings["stat_seconds"] += time.perf_counter() - stat_started

            cached = manifest.get(path_key)
            if cached and float(cached.get("mtime", 0)) == mtime and int(cached.get("size", -1)) == size:
                skipped_files += 1
                report_index_progress(
                    progress_callback,
                    phase="skip",
                    scanned_files=len(current_files),
                    indexed_files=indexed_files,
                    deleted_files=deleted_files,
                    indexed_chunks=indexed_chunks,
                    skipped_files=skipped_files,
                    current_file=path_key,
                    current_index=file_number,
                    timings=round_timings(timings),
                )
                continue

            hash_started = time.perf_counter()
            file_hash = sha256_file(path)
            timings["hash_seconds"] += time.perf_counter() - hash_started
            if cached and cached.get("file_hash") == file_hash:
                cached["mtime"] = mtime
                cached["size"] = size
                manifest[path_key] = cached
                skipped_files += 1
                report_index_progress(
                    progress_callback,
                    phase="skip",
                    scanned_files=len(current_files),
                    indexed_files=indexed_files,
                    deleted_files=deleted_files,
                    indexed_chunks=indexed_chunks,
                    skipped_files=skipped_files,
                    current_file=path_key,
                    current_index=file_number,
                    timings=round_timings(timings),
                )
                continue

            delete_started = time.perf_counter()
            self.store.delete_file(str(path))
            timings["delete_seconds"] += time.perf_counter() - delete_started

            read_started = time.perf_counter()
            text = read_supported_file(path)
            timings["read_seconds"] += time.perf_counter() - read_started
            if not text.strip():
                manifest[path_key] = {
                    "mtime": mtime,
                    "size": size,
                    "file_hash": file_hash,
                    "indexed": False,
                }
                skipped_files += 1
                report_index_progress(
                    progress_callback,
                    phase="read",
                    scanned_files=len(current_files),
                    indexed_files=indexed_files,
                    deleted_files=deleted_files,
                    indexed_chunks=indexed_chunks,
                    skipped_files=skipped_files,
                    current_file=path_key,
                    current_index=file_number,
                    timings=round_timings(timings),
                )
                continue

            chunk_started = time.perf_counter()
            chunks = build_chunks(path, text, file_hash, mtime)
            timings["chunk_seconds"] += time.perf_counter() - chunk_started
            for start in range(0, len(chunks), INDEX_BATCH_SIZE):
                batch = chunks[start : start + INDEX_BATCH_SIZE]
                embed_started = time.perf_counter()
                points = self.store.build_points(batch)
                timings["embed_seconds"] += time.perf_counter() - embed_started

                write_started = time.perf_counter()
                self.store.write_points(points)
                timings["write_seconds"] += time.perf_counter() - write_started
            manifest[path_key] = {
                "mtime": mtime,
                "size": size,
                "file_hash": file_hash,
                "indexed": True,
                "chunks": len(chunks),
            }
            indexed_files += 1
            indexed_chunks += len(chunks)

            report_index_progress(
                progress_callback,
                phase="write",
                scanned_files=len(current_files),
                indexed_files=indexed_files,
                deleted_files=deleted_files,
                indexed_chunks=indexed_chunks,
                skipped_files=skipped_files,
                current_file=path_key,
                current_index=file_number,
                timings=round_timings(timings),
            )

        manifest_save_started = time.perf_counter()
        save_manifest(manifest)
        timings["manifest_save_seconds"] += time.perf_counter() - manifest_save_started
        timings["total_seconds"] = time.perf_counter() - total_started

        result = IndexResponse(
            status="completed",
            indexed_files=indexed_files,
            deleted_files=deleted_files,
            indexed_chunks=indexed_chunks,
            scanned_files=len(current_files),
            skipped_files=skipped_files,
            started_at=started_at,
            finished_at=time.time(),
            timings=round_timings(timings),
        )
        report_index_progress(
            progress_callback,
            phase="completed",
            scanned_files=len(current_files),
            indexed_files=indexed_files,
            deleted_files=deleted_files,
            indexed_chunks=indexed_chunks,
            skipped_files=skipped_files,
            current_file=None,
            current_index=len(current_files),
            timings=result.timings,
        )
        return result

    def iter_files(self) -> Iterable[Path]:
        if not KB_ROOT.exists():
            return []
        files = []
        for path in KB_ROOT.rglob("*"):
            if not path.is_file():
                continue
            if path.suffix.lower() not in KB_EXTENSIONS:
                continue
            normalized = path.as_posix().lower()
            if any(pattern.lower() in normalized for pattern in KB_EXCLUDE_PATTERNS):
                continue
            files.append(path)
        return files


embedder: EmbeddingService | None = None
store: VectorStore | None = None
indexer: FileIndexer | None = None
chat_service: ChatService | None = None
index_jobs: dict[str, IndexJobStatus] = {}
index_jobs_lock = threading.Lock()
index_run_lock = threading.Lock()


def get_embedder() -> EmbeddingService:
    global embedder
    if embedder is None:
        embedder = EmbeddingService()
    return embedder


def get_store() -> VectorStore:
    global store
    if store is None:
        store = VectorStore(get_embedder())
    return store


def get_indexer() -> FileIndexer:
    global indexer
    if indexer is None:
        indexer = FileIndexer(get_store())
    return indexer


def get_chat_service() -> ChatService:
    global chat_service
    if chat_service is None:
        chat_service = ChatService()
    return chat_service


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def load_manifest() -> dict[str, dict[str, Any]]:
    if not MANIFEST_PATH.exists():
        return {}
    try:
        return json.loads(MANIFEST_PATH.read_text(encoding="utf-8"))
    except Exception:
        return {}


def save_manifest(manifest: dict[str, dict[str, Any]]) -> None:
    STATE_DIR.mkdir(parents=True, exist_ok=True)
    MANIFEST_PATH.write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2, sort_keys=True),
        encoding="utf-8",
    )


def job_status_path(job_id: str) -> Path:
    return JOBS_DIR / f"{job_id}.json"


def save_job_status(job: IndexJobStatus) -> None:
    JOBS_DIR.mkdir(parents=True, exist_ok=True)
    job_status_path(job.job_id).write_text(
        job.model_dump_json(indent=2),
        encoding="utf-8",
    )


def load_job_status(job_id: str) -> IndexJobStatus | None:
    path = job_status_path(job_id)
    if not path.exists():
        return None
    try:
        return IndexJobStatus.model_validate_json(path.read_text(encoding="utf-8"))
    except Exception:
        return None


def load_job_statuses() -> list[IndexJobStatus]:
    if not JOBS_DIR.exists():
        return []
    jobs = []
    for path in sorted(JOBS_DIR.glob("*.json")):
        try:
            jobs.append(IndexJobStatus.model_validate_json(path.read_text(encoding="utf-8")))
        except Exception:
            continue
    return jobs


def init_timings() -> dict[str, float]:
    return {
        "scan_seconds": 0.0,
        "manifest_load_seconds": 0.0,
        "stat_seconds": 0.0,
        "hash_seconds": 0.0,
        "delete_seconds": 0.0,
        "read_seconds": 0.0,
        "chunk_seconds": 0.0,
        "embed_seconds": 0.0,
        "write_seconds": 0.0,
        "manifest_save_seconds": 0.0,
        "total_seconds": 0.0,
    }


def round_timings(timings: dict[str, float]) -> dict[str, float]:
    return {key: round(value, 6) for key, value in timings.items()}


def report_index_progress(progress_callback: Any, **progress: Any) -> None:
    if progress_callback is None:
        return
    progress_callback(progress)


def list_index_jobs() -> list[IndexJobStatus]:
    with index_jobs_lock:
        jobs = {job.job_id: job for job in index_jobs.values()}
    for job in load_job_statuses():
        jobs[job.job_id] = job
    return sorted(jobs.values(), key=lambda job: job.started_at or 0.0, reverse=True)


def get_active_index_job() -> IndexJobStatus | None:
    for job in list_index_jobs():
        if job.status == "running":
            return job
    return None


def run_index_job(job_id: str) -> None:
    with index_run_lock:
        with index_jobs_lock:
            job = index_jobs[job_id]
            job.status = "running"
            job.started_at = time.time()
            index_jobs[job_id] = job
            save_job_status(job)

        def progress_callback(progress: dict[str, Any]) -> None:
            with index_jobs_lock:
                current = index_jobs[job_id]
                current.progress = progress
                index_jobs[job_id] = current
                save_job_status(current)

        try:
            result = get_indexer().index(progress_callback=progress_callback)
            with index_jobs_lock:
                index_jobs[job_id] = IndexJobStatus(
                    job_id=job_id,
                    status="completed",
                    started_at=job.started_at,
                    finished_at=time.time(),
                    progress={
                        "phase": "completed",
                        "scanned_files": result.scanned_files,
                        "indexed_files": result.indexed_files,
                        "deleted_files": result.deleted_files,
                        "indexed_chunks": result.indexed_chunks,
                        "skipped_files": result.skipped_files,
                        "timings": result.timings,
                    },
                    result=result,
                )
                save_job_status(index_jobs[job_id])
        except Exception as exc:
            with index_jobs_lock:
                index_jobs[job_id] = IndexJobStatus(
                    job_id=job_id,
                    status="failed",
                    started_at=job.started_at,
                    finished_at=time.time(),
                    error=str(exc),
                    progress=getattr(index_jobs.get(job_id), "progress", None),
                )
                save_job_status(index_jobs[job_id])


def normalize_text(text: str) -> str:
    text = text.replace("\x00", " ")
    text = re.sub(r"\r\n?", "\n", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    text = re.sub(r"[ \t]{2,}", " ", text)
    return text.strip()


def normalize_query(text: str) -> str:
    text = text.replace("\x00", " ")
    text = re.sub(r"[*_`#>~\[\](){}!]+", " ", text)
    text = re.sub(r"\r\n?", "\n", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def normalize_lookup_key(text: str) -> str:
    return re.sub(r"[^\w\u4e00-\u9fff]+", "", text.lower())


def filename_match_score(query_key: str, payload: dict[str, Any]) -> float:
    file_name = str(payload.get("file_name", ""))
    source_path = str(payload.get("source_path", ""))
    source_ref = str(payload.get("source_ref", source_path))

    file_key = normalize_lookup_key(file_name)
    path_key = normalize_lookup_key(source_path)
    ref_key = normalize_lookup_key(source_ref)

    if not query_key:
        return 0.0
    if query_key == file_key:
        return 1.3
    if file_key and query_key in file_key:
        return 1.2
    if query_key == path_key or query_key == ref_key:
        return 1.15
    if query_key in path_key or query_key in ref_key:
        return 1.05
    return 0.0


def merge_matches(vector_matches: list[dict[str, Any]], filename_matches: list[dict[str, Any]]) -> list[dict[str, Any]]:
    merged: dict[tuple[Any, Any], dict[str, Any]] = {}
    for match in [*vector_matches, *filename_matches]:
        key = (match.get("source_ref") or match.get("source_path"), match.get("chunk_index"))
        existing = merged.get(key)
        if existing is None or match.get("score", 0.0) > existing.get("score", 0.0):
            merged[key] = match
    return sorted(merged.values(), key=lambda item: item.get("score", 0.0), reverse=True)


def read_supported_file(path: Path) -> str:
    suffix = path.suffix.lower()
    try:
        if suffix in {
            ".txt",
            ".md",
            ".markdown",
            ".text",
            ".rst",
            ".log",
            ".ini",
            ".cfg",
            ".conf",
            ".yaml",
            ".yml",
            ".json",
            ".xml",
            ".html",
            ".htm",
            ".css",
            ".js",
            ".ts",
            ".tsx",
            ".jsx",
            ".py",
            ".go",
            ".java",
            ".c",
            ".cc",
            ".cpp",
            ".h",
            ".hpp",
            ".rs",
            ".sh",
            ".sql",
        }:
            return normalize_text(path.read_text(encoding="utf-8", errors="ignore"))
        if suffix in {".csv", ".tsv"}:
            return read_delimited_file(path)
        if suffix == ".pdf":
            return read_pdf_file(path)
        if suffix == ".doc":
            return read_doc_file(path)
        if suffix == ".docx":
            document = Document(str(path))
            text = "\n".join(paragraph.text for paragraph in document.paragraphs)
            return normalize_text(text)
        if suffix in {".xls", ".xlsx"}:
            return read_excel_file(path)
    except Exception:
        return ""
    return ""


def read_delimited_file(path: Path) -> str:
    delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
    rows = []
    with path.open("r", encoding="utf-8", errors="ignore", newline="") as handle:
        reader = csv.reader(handle, delimiter=delimiter)
        for row in reader:
            values = [str(cell).strip() for cell in row if str(cell).strip()]
            if values:
                rows.append(" | ".join(values))
    return normalize_text("\n".join(rows))


def read_excel_sheet_texts(path: Path) -> list[dict[str, str]]:
    suffix = path.suffix.lower()
    sheets_text = []
    if suffix == ".xlsx":
        workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
        try:
            for sheet in workbook.worksheets:
                rows = [f"[Sheet] {sheet.title}"]
                rows_included = 0
                for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                    if rows_included >= EXCEL_MAX_ROWS_PER_SHEET:
                        rows.append(f"[Truncated] Reached row limit: {EXCEL_MAX_ROWS_PER_SHEET}")
                        break
                    line = excel_row_to_text(row_idx, row)
                    if line is None:
                        continue
                    rows.append(line)
                    rows_included += 1
                if len(rows) > 1:
                    sheets_text.append({"title": sheet.title, "text": normalize_text("\n".join(rows))})
        finally:
            workbook.close()
        return sheets_text

    import xlrd

    workbook = xlrd.open_workbook(str(path), on_demand=True)
    try:
        for name in workbook.sheet_names():
            sheet = workbook.sheet_by_name(name)
            rows = [f"[Sheet] {name}"]
            rows_included = 0
            for row_idx in range(sheet.nrows):
                if rows_included >= EXCEL_MAX_ROWS_PER_SHEET:
                    rows.append(f"[Truncated] Reached row limit: {EXCEL_MAX_ROWS_PER_SHEET}")
                    break
                row = sheet.row_values(row_idx)
                line = excel_row_to_text(row_idx + 1, row)
                if line is None:
                    continue
                rows.append(line)
                rows_included += 1
            if len(rows) > 1:
                sheets_text.append({"title": name, "text": normalize_text("\n".join(rows))})
        return sheets_text
    finally:
        workbook.release_resources()


def read_excel_file(path: Path) -> str:
    return normalize_text("\n\n".join(sheet["text"] for sheet in read_excel_sheet_texts(path)))


def excel_row_to_text(row_number: int, row: Iterable[Any]) -> str | None:
    values = []
    for column_index, cell in enumerate(row, start=1):
        if column_index > EXCEL_MAX_COLUMNS_PER_ROW:
            values.append(f"[Truncated columns>{EXCEL_MAX_COLUMNS_PER_ROW}]")
            break
        text = excel_cell_to_text(cell)
        if text:
            values.append(f"C{column_index}:{text}")

    if not values:
        if EXCEL_INCLUDE_EMPTY_ROWS:
            return f"R{row_number}: [Empty Row]"
        return None
    return f"R{row_number}: " + " | ".join(values)


def excel_cell_to_text(cell: Any) -> str:
    if cell is None:
        return ""
    if isinstance(cell, str):
        return cell.strip()
    return str(cell).strip()


def read_doc_file(path: Path) -> str:
    antiword_text = run_text_extractor(["antiword", str(path)])
    if antiword_text:
        return antiword_text

    catdoc_text = run_text_extractor(["catdoc", str(path)])
    if catdoc_text:
        return catdoc_text
    return ""


def run_text_extractor(command: list[str]) -> str:
    result = subprocess.run(
        command,
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="ignore",
        check=False,
    )
    if result.returncode == 0 and result.stdout.strip():
        return normalize_text(result.stdout)
    return ""


def read_pdf_file(path: Path) -> str:
    if PDF_EXTRACTOR == "pymupdf" and fitz is not None:
        document = fitz.open(str(path))
        try:
            text = "\n".join(page.get_text("text") for page in document)
        finally:
            document.close()
        return normalize_text(text)

    reader = PdfReader(str(path))
    text = "\n".join(page.extract_text() or "" for page in reader.pages)
    return normalize_text(text)


def split_text(text: str) -> list[str]:
    if len(text) <= CHUNK_SIZE:
        return [text]
    chunks = []
    start = 0
    step = max(1, CHUNK_SIZE - CHUNK_OVERLAP)
    while start < len(text):
        end = min(len(text), start + CHUNK_SIZE)
        chunk = text[start:end].strip()
        if chunk:
            chunks.append(chunk)
        if end >= len(text):
            break
        start += step
    return chunks


def build_chunks(path: Path, text: str, file_hash: str, mtime: float) -> list[Chunk]:
    chunks = []
    rel_name = path.name
    if path.suffix.lower() in {".xls", ".xlsx"}:
        for sheet in read_excel_sheet_texts(path):
            sheet_title = sheet["title"]
            source_ref = f"{path} [Sheet: {sheet_title}]"
            for index, chunk_text in enumerate(split_text(sheet["text"])):
                chunks.append(
                    Chunk(
                        id=str(uuid.uuid5(uuid.NAMESPACE_URL, f"{path}:{sheet_title}:{index}:{file_hash}")),
                        text=chunk_text,
                        payload={
                            "source_path": str(path),
                            "source_ref": source_ref,
                            "sheet_name": sheet_title,
                            "file_name": rel_name,
                            "chunk_index": index,
                            "file_hash": file_hash,
                            "mtime": mtime,
                        },
                    )
                )
        return chunks

    for index, chunk_text in enumerate(split_text(text)):
        chunks.append(
            Chunk(
                id=str(uuid.uuid5(uuid.NAMESPACE_URL, f"{path}:{index}:{file_hash}")),
                text=chunk_text,
                payload={
                    "source_path": str(path),
                    "source_ref": str(path),
                    "file_name": rel_name,
                    "chunk_index": index,
                    "file_hash": file_hash,
                    "mtime": mtime,
                },
            )
        )
    return chunks


def flatten_message_content(content: str | list[dict[str, Any]] | None) -> str:
    if content is None:
        return ""
    if isinstance(content, str):
        return content
    result = []
    for item in content:
        if item.get("type") == "text":
            result.append(item.get("text", ""))
    return "\n".join(part for part in result if part)


def build_openai_response(answer: str, model_name: str) -> dict[str, Any]:
    response_id = f"chatcmpl-{uuid.uuid4().hex}"
    return {
        "id": response_id,
        "object": "chat.completion",
        "created": int(time.time()),
        "model": model_name,
        "choices": [
            {
                "index": 0,
                "message": {"role": "assistant", "content": answer},
                "finish_reason": "stop",
            }
        ],
        "usage": {"prompt_tokens": 0, "completion_tokens": 0, "total_tokens": 0},
    }


def build_streaming_response(answer: str, model_name: str) -> StreamingResponse:
    response_id = f"chatcmpl-{uuid.uuid4().hex}"

    def generate() -> Iterator[str]:
        first_chunk = {
            "id": response_id,
            "object": "chat.completion.chunk",
            "created": int(time.time()),
            "model": model_name,
            "choices": [{"index": 0, "delta": {"role": "assistant"}, "finish_reason": None}],
        }
        yield f"data: {json.dumps(first_chunk, ensure_ascii=False)}\n\n"

        step = 120
        for start in range(0, len(answer), step):
            piece = answer[start : start + step]
            chunk = {
                "id": response_id,
                "object": "chat.completion.chunk",
                "created": int(time.time()),
                "model": model_name,
                "choices": [{"index": 0, "delta": {"content": piece}, "finish_reason": None}],
            }
            yield f"data: {json.dumps(chunk, ensure_ascii=False)}\n\n"

        final_chunk = {
            "id": response_id,
            "object": "chat.completion.chunk",
            "created": int(time.time()),
            "model": model_name,
            "choices": [{"index": 0, "delta": {}, "finish_reason": "stop"}],
        }
        yield f"data: {json.dumps(final_chunk, ensure_ascii=False)}\n\n"
        yield "data: [DONE]\n\n"

    return StreamingResponse(generate(), media_type="text/event-stream")


def proxy_streaming_response(
    stream: Any,
    model_name: str,
    citations: list[str],
) -> StreamingResponse:
    response_id = f"chatcmpl-{uuid.uuid4().hex}"

    def generate() -> Iterator[str]:
        first_chunk = {
            "id": response_id,
            "object": "chat.completion.chunk",
            "created": int(time.time()),
            "model": model_name,
            "choices": [{"index": 0, "delta": {"role": "assistant"}, "finish_reason": None}],
        }
        yield f"data: {json.dumps(first_chunk, ensure_ascii=False)}\n\n"

        for event in stream:
            if not event.choices:
                continue
            choice = event.choices[0]
            delta = getattr(choice, "delta", None)
            content = getattr(delta, "content", None) if delta is not None else None
            finish_reason = getattr(choice, "finish_reason", None)

            if content:
                chunk = {
                    "id": response_id,
                    "object": "chat.completion.chunk",
                    "created": int(time.time()),
                    "model": model_name,
                    "choices": [{"index": 0, "delta": {"content": content}, "finish_reason": None}],
                }
                yield f"data: {json.dumps(chunk, ensure_ascii=False)}\n\n"

            if finish_reason:
                break

        if citations:
            sources_block = "\n\nSources:\n" + "\n".join(dict.fromkeys(citations))
            chunk = {
                "id": response_id,
                "object": "chat.completion.chunk",
                "created": int(time.time()),
                "model": model_name,
                "choices": [{"index": 0, "delta": {"content": sources_block}, "finish_reason": None}],
            }
            yield f"data: {json.dumps(chunk, ensure_ascii=False)}\n\n"

        final_chunk = {
            "id": response_id,
            "object": "chat.completion.chunk",
            "created": int(time.time()),
            "model": model_name,
            "choices": [{"index": 0, "delta": {}, "finish_reason": "stop"}],
        }
        yield f"data: {json.dumps(final_chunk, ensure_ascii=False)}\n\n"
        yield "data: [DONE]\n\n"

    return StreamingResponse(generate(), media_type="text/event-stream")


def build_context(matches: list[dict[str, Any]]) -> str:
    sections = []
    current_size = 0
    for idx, match in enumerate(matches, start=1):
        section = f"[Source {idx}]\nPath: {match.get('source_ref') or match['source_path']}\nChunk: {match['chunk_index']}\nContent:\n{match['text']}"
        if current_size + len(section) > MAX_CONTEXT_CHARS and sections:
            break
        sections.append(section)
        current_size += len(section)
    return "\n\n".join(sections)


def select_matches(matches: list[dict[str, Any]]) -> list[dict[str, Any]]:
    selected = []
    seen = set()
    for match in matches:
        if match.get("score", 0.0) < MIN_RETRIEVAL_SCORE:
            continue
        key = (match.get("source_ref") or match.get("source_path"), match.get("chunk_index"))
        if key in seen:
            continue
        seen.add(key)
        selected.append(match)
    return selected


@app.get("/health")
def health() -> dict[str, str]:
    return {"status": "ok"}


@app.post("/index", response_model=IndexResponse)
def run_index() -> IndexResponse:
    return get_indexer().index()


@app.post("/index/async", response_model=IndexJobStatus)
def run_index_async() -> IndexJobStatus:
    active_job = get_active_index_job()
    if active_job is not None:
        raise HTTPException(status_code=409, detail=f"index job already running: {active_job.job_id}")

    job_id = f"index-{uuid.uuid4().hex}"
    job = IndexJobStatus(job_id=job_id, status="queued")
    with index_jobs_lock:
        index_jobs[job_id] = job
        save_job_status(job)

    thread = threading.Thread(target=run_index_job, args=(job_id,), daemon=True)
    thread.start()
    return job


@app.get("/index/jobs", response_model=list[IndexJobStatus])
def get_index_jobs() -> list[IndexJobStatus]:
    return list_index_jobs()


@app.get("/index/jobs/{job_id}", response_model=IndexJobStatus)
def get_index_job(job_id: str) -> IndexJobStatus:
    with index_jobs_lock:
        job = index_jobs.get(job_id)
    if job is None:
        job = load_job_status(job_id)
    if job is None:
        raise HTTPException(status_code=404, detail="index job not found")
    return job


@app.get("/v1/models")
def list_models() -> dict[str, Any]:
    model_id = CHAT_MODEL or "rag-model"
    return {
        "object": "list",
        "data": [{"id": model_id, "object": "model", "owned_by": "rag-api"}],
    }


@app.post("/v1/chat/completions")
def chat_completions(request: ChatCompletionRequest) -> Any:
    user_messages = [
        {"role": message.role, "content": flatten_message_content(message.content)}
        for message in request.messages
        if message.role in {"system", "user", "assistant"}
    ]
    latest_user = next((msg["content"] for msg in reversed(user_messages) if msg["role"] == "user"), "")
    if not latest_user.strip():
        raise HTTPException(status_code=400, detail="missing user message")

    normalized_query = normalize_query(latest_user)
    search_query = normalized_query or latest_user
    vector_matches = get_store().search(search_query, RETRIEVAL_LIMIT)
    filename_matches = get_store().search_filename_matches(search_query, RETRIEVAL_LIMIT)
    raw_matches = merge_matches(vector_matches, filename_matches)
    matches = select_matches(raw_matches)
    if DEBUG:
        logger.info(
            "retrieval matches query=%s normalized_query=%s vector_matches=%s filename_matches=%s raw_matches=%s selected_matches=%s",
            latest_user,
            search_query,
            json.dumps(vector_matches, ensure_ascii=False),
            json.dumps(filename_matches, ensure_ascii=False),
            json.dumps(raw_matches, ensure_ascii=False),
            json.dumps(matches, ensure_ascii=False),
        )
    context = build_context(matches)

    system_prompt = (
        "You are a retrieval-augmented assistant. Use the provided context first. "
        "If the context is insufficient, say so clearly. Always keep answers concise and include source paths when relevant.\n\n"
        f"Context:\n{context if context else 'No relevant documents were found.'}"
    )
    llm_messages = [{"role": "system", "content": system_prompt}, *user_messages]

    citations = []
    for match in matches:
        citations.append(f"- {match.get('source_ref') or match['source_path']}")

    model_name = request.model or CHAT_MODEL or "rag-model"
    if request.stream:
        stream = get_chat_service().complete_stream(llm_messages, request.temperature or 0.2)
        return proxy_streaming_response(stream, model_name, citations)

    answer = get_chat_service().complete(llm_messages, request.temperature or 0.2)
    if citations:
        answer = f"{answer}\n\nSources:\n" + "\n".join(dict.fromkeys(citations))
    return build_openai_response(answer, model_name)
